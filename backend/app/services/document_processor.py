"""Document processing service."""

import os
import logging
from typing import List, Optional, Union
from pathlib import Path
from io import BytesIO
import re

# PDF & Document processing
try:
    import pdfplumber
    from PyPDF2 import PdfReader
except ImportError:
    logging.warning("PDF libraries not installed")

try:
    from docx import Document as DocxDocument
except ImportError:
    logging.warning("python-docx not installed")

try:
    import openpyxl
except ImportError:
    logging.warning("openpyxl not installed")

logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Process various document types and extract text."""

    ALLOWED_FORMATS = {".pdf", ".docx", ".txt", ".pptx", ".xlsx", ".md"}

    @staticmethod
    def extract_text_from_pdf(file_input: Union[str, BytesIO]) -> str:
        """Extract text from PDF file (supports both file path and BytesIO)."""
        try:
            text = ""
            with pdfplumber.open(file_input) as pdf:
                for page in pdf.pages:
                    text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting PDF: {e}")
            raise

    @staticmethod
    def extract_text_from_docx(file_input: Union[str, BytesIO]) -> str:
        """Extract text from DOCX file (supports both file path and BytesIO)."""
        try:
            doc = DocxDocument(file_input)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text
        except Exception as e:
            logger.error(f"Error extracting DOCX: {e}")
            raise

    @staticmethod
    def extract_text_from_txt(file_input: Union[str, BytesIO]) -> str:
        """Extract text from TXT file (supports both file path and BytesIO)."""
        try:
            if isinstance(file_input, BytesIO):
                content = file_input.read()
                if isinstance(content, bytes):
                    return content.decode("utf-8")
                return content
            else:
                with open(file_input, "r", encoding="utf-8") as f:
                    return f.read()
        except Exception as e:
            logger.error(f"Error reading TXT: {e}")
            raise

    @staticmethod
    def extract_text_from_xlsx(file_input: Union[str, BytesIO]) -> str:
        """Extract text from XLSX file (supports both file path and BytesIO)."""
        try:
            workbook = openpyxl.load_workbook(file_input)
            text = ""
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                text += f"\n=== Sheet: {sheet} ===\n"
                for row in worksheet.iter_rows(values_only=True):
                    text += " | ".join([str(cell) if cell else "" for cell in row]) + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting XLSX: {e}")
            raise

    @staticmethod
    def clean_text(text: str) -> str:
        """Clean and normalize text."""
        # Remove extra whitespace
        text = re.sub(r'\s+', ' ', text)
        # Remove special characters (keep alphanumeric, spaces, and basic punctuation)
        text = re.sub(r'[^\w\s.,!?;:\-()]', '', text)
        # Strip leading/trailing whitespace
        text = text.strip()
        return text

    @staticmethod
    def extract_text(file_input: Union[str, BytesIO], file_ext: Optional[str] = None) -> str:
        """Extract text from any supported file format.
        
        Args:
            file_input: File path (str) or BytesIO object
            file_ext: File extension (e.g., '.pdf'). Auto-detected if not provided
        """
        # Determine file extension
        if file_ext is None:
            if isinstance(file_input, str):
                file_ext = Path(file_input).suffix.lower()
            else:
                raise ValueError("file_ext must be provided when using BytesIO object")

        if file_ext not in DocumentProcessor.ALLOWED_FORMATS:
            raise ValueError(f"Unsupported file format: {file_ext}")

        try:
            if file_ext == ".pdf":
                text = DocumentProcessor.extract_text_from_pdf(file_input)
            elif file_ext == ".docx":
                text = DocumentProcessor.extract_text_from_docx(file_input)
            elif file_ext in [".txt", ".md"]:
                text = DocumentProcessor.extract_text_from_txt(file_input)
            elif file_ext == ".xlsx":
                text = DocumentProcessor.extract_text_from_xlsx(file_input)
            else:
                raise ValueError(f"No extraction method for {file_ext}")

            # Clean the text
            text = DocumentProcessor.clean_text(text)
            return text
        except Exception as e:
            logger.error(f"Error processing document: {e}")
            raise


class TextChunker:
    """Split text into chunks with overlap."""

    def __init__(self, chunk_size: int = 500, overlap: int = 50):
        """Initialize chunker with parameters.
        
        Args:
            chunk_size: Target size for each chunk (in tokens)
            overlap: Number of tokens to overlap between chunks
        """
        self.chunk_size = chunk_size
        self.overlap = overlap

    @staticmethod
    def count_tokens(text: str) -> int:
        """Approximate token count (rough estimate: 1 token ≈ 4 chars)."""
        return len(text) // 4

    def chunk_text(self, text: str, min_chunk_size: int = 100) -> List[dict]:
        """Split text into overlapping chunks with metadata.
        
        Returns:
            List of dicts with 'text' and 'tokens' keys
        """
        sentences = text.split(". ")
        chunks = []
        current_chunk = ""

        for sentence in sentences:
            sentence = sentence.strip() + ". "
            sentence_tokens = self.count_tokens(sentence)

            # If adding this sentence would exceed chunk size
            if self.count_tokens(current_chunk) + sentence_tokens > self.chunk_size:
                if current_chunk.strip():
                    chunk_tokens = self.count_tokens(current_chunk)
                    chunks.append({
                        'text': current_chunk.strip(),
                        'tokens': chunk_tokens
                    })
                # Start new chunk with overlap from previous
                overlap_text = " ".join(current_chunk.split()[-self.overlap:]) if self.overlap > 0 else ""
                current_chunk = overlap_text + " " + sentence if overlap_text else sentence
            else:
                current_chunk += sentence

        # Add remaining chunk
        if current_chunk.strip() and self.count_tokens(current_chunk) > min_chunk_size:
            chunk_tokens = self.count_tokens(current_chunk)
            chunks.append({
                'text': current_chunk.strip(),
                'tokens': chunk_tokens
            })

        return chunks
